import datetime

from django.db import models
from django.db.models import Sum, F
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from common.enums import Status
from core.permissions import (
    StaffIsAdmin,
    StaffIsFrontDeskProductReturn,
    StaffIsTelemarketer,
    StaffIsDistributionT3,
    CheckAnyPermission,
)
from ecommerce.enums import ShortReturnLogType
from ecommerce.filters import ShortReturnItemFilter
from ecommerce.models import ShortReturnItem


class ReturnReportList(APIView):
    available_permission_classes = (
        StaffIsAdmin,
        StaffIsFrontDeskProductReturn,
        StaffIsTelemarketer,
        StaffIsDistributionT3,
    )
    permission_classes = (CheckAnyPermission,)
    filterset_class = ShortReturnItemFilter

    def get_queryset(self):
        return_logs = ShortReturnItem.objects.filter(
            type=ShortReturnLogType.RETURN,
            status=Status.ACTIVE,
            short_return_log__approved_by__isnull=False
        ).annotate(
            total_quantity=Sum('quantity'),
            approved_by=models.functions.Concat(
                models.functions.Cast('short_return_log__approved_by__first_name', models.CharField()),
                models.Value(' '),
                models.functions.Cast('short_return_log__approved_by__last_name', models.CharField()),
                models.Value('['),
                models.functions.Cast('short_return_log__approved_by__id', models.CharField()),
                models.Value(']'),
                models.Value(':'),
                models.functions.Cast('quantity', models.CharField()),
                delimiter=', '
            ),
            approved_quantity=F('quantity')
        ).values(
            'product_name', 'stock', 'total_quantity', 'approved_by', 'approved_quantity', 'stock__product_full_name'
        )

        return return_logs

    def get(self, request):
        import pandas as pd
        import re
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Alignment

        return_logs = self.filterset_class(request.GET, queryset=self.get_queryset()).qs
        if return_logs.exists():
            df = pd.DataFrame.from_records(return_logs)
            df = df.explode('approved_by')
            df[['approved_by', 'approved_quantity']] = df['approved_by'].str.split(':', expand=True)
            df['approved_quantity'] = pd.to_numeric(df['approved_quantity'], errors='coerce')
            df['approved_quantity'] = df['approved_quantity'].astype(int)
            grouped = df.groupby(['product_name', 'approved_by'])['approved_quantity'].sum().reset_index()
            pivot_table = grouped.pivot(index='product_name', columns='approved_by', values='approved_quantity')
            pivot_table.insert(0, 'Stock', df.groupby('product_name')['stock'].first())
            pivot_table.insert(0, 'full_name', df.groupby('product_name')['stock__product_full_name'].first())
            pivot_table = pivot_table.sort_values(by='full_name', ascending=True)
            pivot_table = pivot_table.drop('full_name', axis=1)
            pivot_table['Total Qty'] = pd.to_numeric(df.groupby('product_name')['total_quantity'].sum())
            pivot_table.reset_index(inplace=True)
            pivot_table.insert(0, 'SL', range(1, 1 + len(pivot_table)))
            total_row = pivot_table[pivot_table.columns[3:]].sum(numeric_only=True, axis=0)
            total_row["SL"] = "Total"
            pivot_table = pd.concat([pivot_table, total_row.to_frame().T], ignore_index=True)
            pivot_table.iloc[:, 3:] = pivot_table.iloc[:, 3:].fillna(0)
            pivot_table.rename(columns={'product_name': 'Item'}, inplace=True)

            # make multi header for pivot table
            approved_by_list = pivot_table.columns[3:-1].tolist()
            header_list = [' ', ' ', ' '] + ["Approved By"] * len(approved_by_list) + [' ']
            new_header_list = ['SL', 'Item', 'Stock'] + approved_by_list + ['Total Qty']
            new_table_header_list = [header_list, new_header_list]
            pivot_table.columns = new_table_header_list
            pivot_table.set_index(pivot_table.columns[0], inplace=True)
            pivot_table.index.rename('SL', inplace=True)

            # make Excel file
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.append(header_list)
            worksheet.append(new_header_list)

            # add the data to the worksheet
            for r in dataframe_to_rows(pivot_table, index=True, header=False):
                worksheet.append(r)

            # set column width
            for i in range(pivot_table.shape[1]):
                worksheet.column_dimensions[get_column_letter(i + 1)].width = 20

            # merge cells for multi header and set alignment
            worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=worksheet.max_column - 1)
            worksheet.row_dimensions[1].alignment = Alignment(horizontal="center", vertical="center")
            worksheet.row_dimensions[2].alignment = Alignment(horizontal="center", vertical="center")
            # set column width for first two columns
            worksheet.column_dimensions['A'].width = 10
            worksheet.column_dimensions['B'].width = 40
            # delete default empty row from pivot table
            worksheet.delete_rows(3)
            # remove "[,number,]" from second row
            for i in range(3, worksheet.max_column + 1):
                cell = worksheet.cell(row=2, column=i)
                cell.value = re.sub(r'[\[\]\d+]', '', cell.value)

            # Write pivot table to Excel file
            today = datetime.datetime.now()
            filename = 'short_return_report_{}.xlsx'.format(today)
            response = HttpResponse(content_type='text/xlsx')
            response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
            workbook.save(response)

            return response
        else:
            return Response(status=status.HTTP_204_NO_CONTENT)
